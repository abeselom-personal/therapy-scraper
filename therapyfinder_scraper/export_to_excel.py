import logging
import os
from datetime import datetime
from typing import Any, Dict, List, Optional

import pandas as pd
from pymongo import MongoClient
from pymongo.errors import ConnectionFailure, OperationFailure

# Configure logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

# MongoDB connection configuration
MONGO_HOST = os.getenv("MONGO_HOST", "mongodb")
MONGO_PORT = int(os.getenv("MONGO_PORT", "27017"))
MONGO_DB = os.getenv("MONGO_DB", "therapyfinder_speed_test")
MONGO_USER = os.getenv("MONGO_USER", "scraper")
MONGO_PASSWORD = os.getenv("MONGO_PASSWORD", "scraper")

OUTPUT_DIR = "/app/exports/therapyfinder/"


# Analysis counters
class AnalysisCounters:
    def __init__(self):
        self.total_clinicians = 0
        self.successful_flattens = 0
        self.failed_flattens = 0
        self.duplicates_found = 0
        self.records_with_missing_data = 0
        self.fields_with_empty_data = {}
        self.records_exported = 0

    def increment_field_empty(self, field_name: str):
        self.fields_with_empty_data[field_name] = (
            self.fields_with_empty_data.get(field_name, 0) + 1
        )

    def print_summary(self):
        logger.info("=== DATA ANALYSIS SUMMARY ===")
        logger.info(f"Total clinicians in database: {self.total_clinicians}")
        logger.info(f"Successfully processed: {self.successful_flattens}")
        logger.info(f"Failed to process: {self.failed_flattens}")
        logger.info(f"Duplicates removed: {self.duplicates_found}")
        logger.info(
            f"Records with missing data: {self.records_with_missing_data}"
        )
        logger.info(f"Final records exported: {self.records_exported}")

        if self.fields_with_empty_data:
            logger.info("=== EMPTY FIELD ANALYSIS ===")
            for field, count in sorted(
                self.fields_with_empty_data.items(),
                key=lambda x: x[1],
                reverse=True,
            )[:10]:
                percentage = (
                    (count / self.successful_flattens) * 100
                    if self.successful_flattens > 0
                    else 0
                )
                logger.info(f"  {field}: {count} ({percentage:.1f}%)")


def get_mongo_client() -> MongoClient:
    """
    Establish MongoDB connection with error handling
    """
    try:
        connection_string = f"mongodb://{MONGO_USER}:{MONGO_PASSWORD}@{MONGO_HOST}:{MONGO_PORT}/{MONGO_DB}?authSource=admin"
        client = MongoClient(
            connection_string,
            serverSelectionTimeoutMS=5000,  # 5 second timeout
            connectTimeoutMS=10000,
            socketTimeoutMS=30000,
        )
        # Test the connection
        client.admin.command("ping")
        logger.info("Successfully connected to MongoDB")
        return client
    except ConnectionFailure as e:
        logger.error(f"Failed to connect to MongoDB: {e}")
        raise
    except Exception as e:
        logger.error(f"Unexpected error connecting to MongoDB: {e}")
        raise


def safe_get_list(data: Dict, key: str, default: List = None) -> List:
    """
    Safely extract list from dictionary with comprehensive type checking
    """
    if default is None:
        default = []

    try:
        value = data.get(key)
        if value is None:
            return default
        elif isinstance(value, list):
            return value
        elif isinstance(value, str):
            # Try to handle string representations of lists
            return [value]
        else:
            logger.warning(
                f"Unexpected type for key '{key}': {type(value)}. Returning default."
            )
            return default
    except Exception as e:
        logger.warning(f"Error extracting list for key '{key}': {e}")
        return default


def safe_join_strings(
    data_list: List, separator: str = ", ", field_name: str = "unknown"
) -> str:
    """
    Safely join strings from a list of dictionaries with enhanced error handling
    """
    if not data_list:
        return ""

    try:
        joined_string = separator.join(
            [
                str(item.get("name", ""))
                for item in data_list
                if item and isinstance(item, dict) and item.get("name")
            ]
        )
        return joined_string
    except Exception as e:
        logger.warning(f"Error joining strings for field '{field_name}': {e}")
        return ""


def analyze_data_quality(clinician: Dict, counters: AnalysisCounters) -> None:
    """
    Analyze data quality and update counters
    """
    attributes = clinician.get("attributes", {})

    # Check for missing critical fields
    critical_fields = ["display_name", "links"]
    missing_critical = any(
        not clinician.get(field) for field in critical_fields
    )
    if missing_critical:
        counters.records_with_missing_data += 1

    # Analyze empty fields in the flattened structure (predicted)
    test_flattened = flatten_clinician_data(
        clinician, counters, analyze_only=True
    )

    for field, value in test_flattened.items():
        if not value and value != 0:  # 0 might be valid for some numeric fields
            counters.increment_field_empty(field)


def flatten_clinician_data(
    clinician: Dict, counters: AnalysisCounters, analyze_only: bool = False
) -> Dict[str, Any]:
    """
    Flatten clinician data with comprehensive error handling and data analysis

    Args:
        clinician: Raw clinician data from MongoDB
        counters: Analysis counters instance
        analyze_only: If True, only analyze don't count as successful processing
    """
    try:
        attributes = clinician.get("attributes", {})

        # Extract fees safely
        fees = attributes.get("fees", [{}])
        min_fee = ""
        max_fee = ""
        if fees and isinstance(fees, list) and fees[0]:
            min_fee = fees[0].get("min_fee", "")
            max_fee = fees[0].get("max_fee", "")

        flattened = {
            "clinician_id": clinician.get("clinician_id", ""),
            "Url": f"https://therapyfinder.com/therapist/{attributes.get('slug', '')}",
            "Name": clinician.get("display_name", ""),
            "NPI": attributes.get("npiNumber", ""),
            "Profession": attributes.get("title", ""),
            "Clinic Name": "",
            "Bio": clinician.get("bio", ""),
            "Additional Focus Areas": safe_join_strings(
                safe_get_list(attributes, "allSpecialties"),
                field_name="Additional Focus Areas",
            ),
            "Treatment Approaches": safe_join_strings(
                safe_get_list(attributes, "allServices"),
                field_name="Treatment Approaches",
            ),
            "Appointment Types": safe_join_strings(
                safe_get_list(attributes, "treatmentTypes"),
                field_name="Appointment Types",
            ),
            "Communities": safe_join_strings(
                safe_get_list(attributes, "communities"),
                field_name="Communities",
            ),
            "Age Groups": safe_join_strings(
                safe_get_list(attributes, "ageGroups"), field_name="Age Groups"
            ),
            "Languages": safe_join_strings(
                safe_get_list(attributes, "languages"), field_name="Languages"
            ),
            "Highlights": "",
            "Gender": safe_join_strings(
                safe_get_list(attributes, "gender"), field_name="Gender"
            ),
            "Pronouns": safe_join_strings(
                safe_get_list(attributes, "pronouns"), field_name="Pronouns"
            ),
            "Race Ethnicity": safe_join_strings(
                safe_get_list(attributes, "raceEthnicities"),
                field_name="Race Ethnicity",
            ),
            "Licenses": safe_join_strings(
                safe_get_list(
                    attributes, "allInsuranceCarriers"
                ),  # Note: This might be misnamed
                field_name="Licenses",
            ),
            "Locations": f"{clinician.get('location_city', '')}, {clinician.get('location_state', '')}".strip(
                ", "
            ),
            "Education": "",
            "Faiths": safe_join_strings(
                safe_get_list(attributes, "faiths"), field_name="Faiths"
            ),
            "Min Session Price": min_fee,
            "Max Session Price": max_fee,
            "Pay Out Of Pocket Status": "",
            "Individual Service Rates": safe_join_strings(
                safe_get_list(
                    attributes, "allServices"
                ),  # Duplicate of Treatment Approaches?
                field_name="Individual Service Rates",
            ),
            "General Payment Options": "",
            "Booking Summary": "",
            "Booking Url": "",
            "Listed In States": clinician.get("location_state", ""),
            "States": clinician.get("location_state", ""),
            "Listed In Websites": "",
            "Urls": attributes.get("profileImgUrl", ""),
            "Connect Link - Facebook": attributes.get("facebookUrl", ""),
            "Connect Link - Instagram": attributes.get("instagramUrl", ""),
            "Connect Link - LinkedIn": attributes.get("linkedinUrl", ""),
            "Connect Link - Twitter": attributes.get("twitterUrl", ""),
            "Connect Link - Website": "",
            "Main Specialties": safe_join_strings(
                safe_get_list(
                    attributes, "allSpecialties"
                ),  # Duplicate of Additional Focus Areas?
                field_name="Main Specialties",
            ),
            "Accepted IPs": safe_join_strings(
                safe_get_list(attributes, "allInsuranceCarriers"),
                field_name="Accepted IPs",
            ),
            "Sr. NO": "",
        }

        # Clean None values
        for key in flattened:
            if flattened[key] is None:
                flattened[key] = ""

        if not analyze_only:
            counters.successful_flattens += 1

        return flattened

    except Exception as e:
        if not analyze_only:
            counters.failed_flattens += 1
            logger.error(f"Error flattening clinician data: {e}")
            logger.debug(
                f"Problematic clinician: {clinician.get('display_name', 'Unknown')}"
            )

        # Return minimal structure even on failure
        return {
            "Url": clinician.get("links", {}).get("self", ""),
            "Name": "ERROR_PROCESSING_DATA",
        }


def export_clinicians_to_excel():
    """
    Export clinicians to Excel with detailed logging and memory efficiency.
    """
    import os
    from datetime import datetime

    import pandas as pd
    from openpyxl import Workbook, load_workbook
    from pymongo.errors import ConnectionFailure, OperationFailure

    counters = AnalysisCounters()
    failed_records = []

    try:
        client = get_mongo_client()
        db = client[MONGO_DB]
        logger.info(f"Connected to database: {MONGO_DB}")

        cursor = db.clinicians.find(batch_size=500)
        total_clinicians = db.clinicians.count_documents({})
        counters.total_clinicians = total_clinicians
        logger.info(f"Total clinicians in database: {total_clinicians}")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join(
            OUTPUT_DIR, f"therapyfinder_clinicians_{timestamp}.xlsx"
        )
        os.makedirs(OUTPUT_DIR, exist_ok=True)

        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title="therapyfinder")
        headers_written = False
        row_number = 1

        for i, clinician in enumerate(cursor, 1):
            if i % 500 == 0:
                logger.info(f"Processed {i} clinicians...")

            try:
                analyze_data_quality(clinician, counters)
                flat = flatten_clinician_data(
                    clinician, counters, analyze_only=False
                )
                flat["Sr. NO"] = i

                if not headers_written:
                    ws.append(list(flat.keys()))
                    headers_written = True

                ws.append(list(flat.values()))
            except Exception as e:
                failed_records.append(
                    {
                        "index": i,
                        "name": clinician.get("display_name", "Unknown"),
                        "error": str(e),
                    }
                )
                counters.failed_flattens += 1
                logger.warning(f"Failed processing clinician {i}: {e}")
                continue

        wb.save(output_file)
        counters.records_exported = i - counters.failed_flattens
        logger.info(f"Excel exported successfully: {output_file}")

        # Print summary
        counters.print_summary()
        if failed_records:
            logger.warning(f"Total failed records: {len(failed_records)}")
            for fr in failed_records[:5]:
                logger.warning(
                    f"Record {fr['index']} ({fr['name']}): {fr['error']}"
                )
            if len(failed_records) > 5:
                logger.warning(
                    f"...and {len(failed_records) - 5} more failed records."
                )

        # Export to collection.xlsx safely
        collection_file = "/app/exports/collection.xlsx"
        website_name = "therapyfinder"

        if os.path.exists(collection_file):
            book = load_workbook(collection_file)
            if website_name in book.sheetnames:
                std = book[website_name]
                book.remove(std)
            with pd.ExcelWriter(
                collection_file, engine="openpyxl", mode="a"
            ) as writer:
                df = pd.DataFrame(list(db.clinicians.find()))
                df.to_excel(writer, sheet_name=website_name, index=False)
                analysis_df = pd.DataFrame({f"Total {website_name}": [len(df)]})
                analysis_df.to_excel(writer, sheet_name="analysis", index=False)
        else:
            with pd.ExcelWriter(collection_file, engine="openpyxl") as writer:
                df = pd.DataFrame(list(db.clinicians.find()))
                df.to_excel(writer, sheet_name=website_name, index=False)
                analysis_df = pd.DataFrame({f"Total {website_name}": [len(df)]})
                analysis_df.to_excel(writer, sheet_name="analysis", index=False)

    except ConnectionFailure as e:
        logger.error(f"Database connection failed: {e}")
        raise
    except OperationFailure as e:
        logger.error(f"Database operation failed: {e}")
        raise
    except Exception as e:
        logger.error(f"Unexpected error during export: {e}")
        raise
    finally:
        try:
            client.close()
            logger.info("Database connection closed")
        except:
            pass


if __name__ == "__main__":
    try:
        export_clinicians_to_excel()
    except Exception as e:
        logger.error(f"Application failed: {e}")
        exit(1)
